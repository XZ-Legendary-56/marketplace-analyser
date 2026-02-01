# download_process/parsers/analysis_goods_script.py
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import os
import json
import sys
import numpy as np
import re
from openpyxl import load_workbook
from statsmodels.tsa.statespace.sarimax import SARIMAX
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from sklearn.metrics import mean_absolute_error
import warnings

warnings.filterwarnings("ignore")



def parse_data(file_path, table_name, start_text, id_prefixes):
    if not os.path.exists(file_path): return []
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    data, header_found = [], False
    for row in sheet.iter_rows(values_only=True):
        if not row: continue
        search_cell = str(row[0] if table_name == "Продаж" else row[5])
        if start_text in search_cell:
            header_found = True; continue
        if header_found:
            id_cell = str(row[0] if table_name == "Продаж" else row[5])
            if any(id_cell.startswith(p) for p in id_prefixes):
                try:
                    if table_name == "Продаж":
                        data.append({"ID": row[0], "Название": row[1], "Выручка": float(row[2] or 0), "Дата": row[3]})
                    else: # Услуги
                        data.append({"Услуга": row[6], "Цена": float(str(row[8] or 0).replace(",", "."))})
                except (ValueError, IndexError): continue
    return data

def parse_cost_data(file_path):
    if not os.path.exists(file_path): return {}
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    cost_data, header_found, header_passed = {}, False, False
    for row in sheet.iter_rows(values_only=True):
        if not row: continue
        if "Таблица себестоимости" in str(row[0]):
            header_found = True; continue
        if header_found:
            if not header_passed:
                header_passed = True; continue
            if row[0] and row[1]:
                try: cost_data[row[0]] = float(row[1])
                except (ValueError, TypeError): cost_data[row[0]] = 0
    return cost_data

def build_sales_dict(sales_data):
    d = {}
    for sale in sales_data:
        p = sale["Название"]; d.setdefault(p, []).append([sale["Выручка"], sale["Дата"]])
    return d

def build_services_dict(services_data):
    d = {}
    for service in services_data:
        d.setdefault(service['Услуга'], []).append(service['Цена'])
    return d

def calculate_net_profit(sales_dict, services_dict, product_name, cost_dict):
    if product_name not in sales_dict: return 0
    revenue = sum(s[0] for s in sales_dict[product_name])
    total_revenue_all_products = sum(s[0] for p_sales in sales_dict.values() for s in p_sales)
    share = revenue / total_revenue_all_products if total_revenue_all_products > 0 else 0
    total_services_cost = sum(sum(c) for c in services_dict.values()) # Теперь это работает
    distributed_cost = total_services_cost * share
    total_cost = cost_dict.get(product_name, 0) * len(sales_dict[product_name])
    return revenue - total_cost - distributed_cost

def plot_pie_chart(data, labels, title, filename, graphs_path):
    pos_data, pos_labels = zip(*[(d, l) for d, l in zip(data, labels) if d > 0]) if any(d > 0 for d in data) else ([], [])
    if not pos_data: return
    plt.figure(figsize=(8, 6))
    plt.pie(pos_data, labels=pos_labels, autopct='%1.1f%%', startangle=90, colors=['#ff9999', '#66b3ff', '#99ff99'])
    plt.title(title, pad=20); plt.savefig(os.path.join(graphs_path, filename), bbox_inches='tight', dpi=100); plt.close()

def plot_trend(series, marketplace_name, product_name, y_label, chart_type, graphs_path, filename_prefix):
    if len(series) < 2: return
    df = series.to_frame('value').reset_index()
    x, y = np.arange(len(df)), df['value'].values
    trend = np.polyfit(x, y, 1); trend_line = np.poly1d(trend)(x)
    trend_color = 'green' if trend[0] > 0 else 'red'
    plt.figure(figsize=(10, 5))
    if chart_type == 'line': plt.plot(df['date'], y, marker='o', label=y_label)
    else: plt.bar(df['date'], y, color='skyblue', label=y_label)
    plt.plot(df['date'], trend_line, '--', color=trend_color, label='Тренд')
    plt.title(f'Динамика {y_label.lower()} на {marketplace_name}\nТовар: {product_name}')
    plt.xlabel('Дата'); plt.ylabel(y_label); plt.legend(); plt.grid(True)
    plt.xticks(rotation=45); plt.tight_layout()
    plt.savefig(os.path.join(graphs_path, f'{filename_prefix}_{marketplace_name.lower().replace(" ", "_")}.png'), dpi=100); plt.close()

def forecast_revenue(series, method):
    if len(series) < 14: method = 'mean'
    if method == 'mean':
        non_zero = series[series > 0]
        if non_zero.empty: return np.zeros(30)
        avg, std = non_zero.mean(), non_zero.std()
        active_days_ratio = len(non_zero) / len(series)
        expected_active_days = int(30 * active_days_ratio)
        forecast = np.zeros(30)
        active_indices = np.random.choice(30, expected_active_days, replace=False)
        forecast[active_indices] = np.random.normal(avg, std / 2, expected_active_days)
        return np.maximum(0, forecast)
    try:
        if method == 'linear': model = np.poly1d(np.polyfit(np.arange(len(series)), series, 1)); return np.maximum(0, model(np.arange(len(series), len(series) + 30)))
        if method == 'holt': model = ExponentialSmoothing(series, trend='add').fit(); return np.maximum(0, model.forecast(30))
        if method == 'sarima': model = SARIMAX(series, order=(1, 1, 1), seasonal_order=(1, 1, 1, 7)).fit(disp=False); return np.maximum(0, model.forecast(30))
    except: return forecast_revenue(series, 'holt') # Fallback

def best_forecast(series, y_label, graphs_path):
    if series.empty or len(series) < 2: return [], None
    methods = ['mean', 'linear', 'holt', 'sarima']
    best_method, best_mae = 'mean', float('inf')
    train_size = int(len(series) * 0.8)
    if train_size < 2: return forecast_revenue(series, 'mean'), 'mean' # Not enough data for split
    train, test = series[:train_size], series[train_size:]
    for method in methods:
        try:
            forecast = forecast_revenue(train, method)[:len(test)]; mae = mean_absolute_error(test, forecast)
            if mae < best_mae: best_mae, best_method = mae, method
        except: continue
    forecast = forecast_revenue(series, best_method)
    plt.figure(figsize=(12, 6)); plt.plot(series.index, series, label='История'); future_dates = pd.date_range(series.index[-1] + pd.Timedelta(days=1), periods=30)
    plt.plot(future_dates, forecast, 'r--', label=f'Прогноз ({best_method})'); plt.title(f'Прогноз {y_label.lower()} на следующий месяц'); plt.legend(); plt.grid()
    plt.savefig(os.path.join(graphs_path, f'forecast_{series.name.lower().replace(" ","_")}.png')); plt.close()
    return forecast, best_method


def main(input_path, output_path, product_name, relative_json_path):
    try:
        # --- ПУТИ И ЧТЕНИЕ ДАННЫХ ---
        graphs_path = os.path.join(output_path, 'graphs')
        os.makedirs(graphs_path, exist_ok=True)
        cost_dict = parse_cost_data(os.path.join(input_path, "costs.xlsx"))
        
        marketplaces = { "Wildberries": "wb", "Ozon": "ozon", "Яндекс.Маркет": "ym"}
        mp_data = {}

        for mp_name, prefix in marketplaces.items():
            sales = parse_data(os.path.join(input_path, f"{prefix}_processed_report.xlsx"), "Продаж", "Таблица Продаж", ["WB-", "oz-", "YAN"])
            services = parse_data(os.path.join(input_path, f"{prefix}_processed_report.xlsx"), "Услуг", "Таблица затрат на услуги", ["WB-", "SRV", "YAN"])
            
            # ИСПРАВЛЕНИЕ: Преобразуем список услуг в словарь
            mp_data[mp_name] = {
                "sales_dict": build_sales_dict(sales), 
                "services_dict": build_services_dict(services)
            }


        result = {"product_name": product_name, "marketplaces": {}}
        all_revenues, all_counts, all_profits = [], [], []

        for mp_name, data in mp_data.items():
            sales_dict, services_dict = data['sales_dict'], data['services_dict']
            revenue = sum(s[0] for s in sales_dict.get(product_name, []))
            sales_count = len(sales_dict.get(product_name, []))
            profit = calculate_net_profit(sales_dict, services_dict, product_name, cost_dict)
            
            # Тренды
            daily_series = pd.DataFrame([{'date': s[1], 'revenue': s[0]} for s in sales_dict.get(product_name, [])]).groupby('date')['revenue'].sum().asfreq('D').fillna(0)
            count_series = pd.DataFrame([{'date': s[1], 'count': 1} for s in sales_dict.get(product_name, [])]).groupby('date')['count'].sum().asfreq('D').fillna(0)
            
            plot_trend(daily_series, mp_name, product_name, "Выручка", "line", graphs_path, "revenue_trend")
            plot_trend(count_series, mp_name, product_name, "Количество продаж", "bar", graphs_path, "sales_trend")
            
            # Прогнозы
            daily_series.name = mp_name # Установка имени для файла графика
            count_series.name = mp_name # Установка имени для файла графика
            rev_forecast, rev_method = best_forecast(daily_series, "Выручка", graphs_path)
            count_forecast, count_method = best_forecast(count_series, "Количество продаж", graphs_path)

            result['marketplaces'][mp_name] = {
                "revenue": revenue, "sales_count": sales_count, "profit": profit,
                "forecast": {"revenue": sum(rev_forecast), "sales_count": int(sum(count_forecast)), "method": rev_method},
                "charts": {
                    "revenue_trend": os.path.join(relative_json_path, 'graphs', f'revenue_trend_{mp_name.lower().replace(" ", "_")}.png').replace('\\', '/'),
                    "sales_trend": os.path.join(relative_json_path, 'graphs', f'sales_trend_{mp_name.lower().replace(" ", "_")}.png').replace('\\', '/'),
                    "forecast": os.path.join(relative_json_path, 'graphs', f'forecast_{mp_name.lower().replace(" ","_")}.png').replace('\\', '/') if rev_method else None,
                }
            }
            all_revenues.append(revenue); all_counts.append(sales_count); all_profits.append(profit)
        
        # Общие диаграммы
        mp_names_list = list(marketplaces.keys())
        plot_pie_chart(all_revenues, mp_names_list, f'Распределение выручки\n{product_name}', 'revenue_distribution.png', graphs_path)
        plot_pie_chart(all_counts, mp_names_list, f'Распределение продаж\n{product_name}', 'sales_count_distribution.png', graphs_path)
        plot_pie_chart(all_profits, mp_names_list, f'Распределение прибыли\n{product_name}', 'net_profit_distribution.png', graphs_path)
        

        def get_best_mp(data_key):
            forecasts = [(mp, d['forecast'][data_key]) for mp, d in result['marketplaces'].items() if d['forecast'][data_key] is not None]
            return max(forecasts, key=lambda item: item[1], default=(None, None))

        best_rev_mp, best_rev_val = get_best_mp('revenue')
        best_count_mp, best_count_val = get_best_mp('sales_count')
        
        result["general_stats"] = {
            "best_marketplace_by_revenue": {"name": best_rev_mp, "forecast_revenue": best_rev_val},
            "best_marketplace_by_sales_count": {"name": best_count_mp, "forecast_sales": best_count_val},
            "charts": {
                "revenue_distribution": os.path.join(relative_json_path, 'graphs', 'revenue_distribution.png').replace('\\', '/'),
                "sales_count_distribution": os.path.join(relative_json_path, 'graphs', 'sales_count_distribution.png').replace('\\', '/'),
                "net_profit_distribution": os.path.join(relative_json_path, 'graphs', 'net_profit_distribution.png').replace('\\', '/'),
            }
        }
        

        def convert(o):
            if isinstance(o, (np.int64, np.int32)): return int(o)
            if isinstance(o, (np.float64, np.float32)): return round(float(o), 2)
            if isinstance(o, (pd.Timestamp, datetime)): return o.strftime('%Y-%m-%d %H:%M:%S')
            raise TypeError

        safe_product_name = re.sub(r'[\\/*?:"<>|]', "", product_name)
        json_filename = os.path.join(output_path, f'product_analysis_{safe_product_name}.json')
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=4, default=convert)
        
        print(json.dumps({"status": "success", "message": "Product analysis complete.", "json_file": json_filename}))
    
    except Exception as e:
        import traceback
        exc_type, exc_value, exc_traceback = sys.exc_info()
        error_details = traceback.format_exception(exc_type, exc_value, exc_traceback)
        print(json.dumps({"status": "error", "message": f"An error occurred: {str(e)}", "details": "".join(error_details)}))

if __name__ == "__main__":
    if len(sys.argv) != 5:
        print(json.dumps({"status": "error", "message": "Invalid number of arguments. Expected: input_path, output_path, product_name, relative_json_path"}))
    else:
        main(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])