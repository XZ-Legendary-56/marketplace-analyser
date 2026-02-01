import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import os
import json
import sys
import numpy as np
from openpyxl import load_workbook
import warnings


warnings.filterwarnings('ignore')


def parse_sales_data(mp_path):
    if not os.path.exists(mp_path): return []
    mp = load_workbook(mp_path, data_only=True)
    sheet = mp.active
    sales_data = []
    found_sales_header = False
    for row in sheet.iter_rows(values_only=True):
        if not row: continue
        if "Таблица Продаж" in str(row[0]):
            found_sales_header = True
            continue
        if found_sales_header:
            if len(row) >= 4 and row[0] and (str(row[0]).startswith("WB-") or str(row[0]).startswith("oz-") or str(row[0]).startswith("YAN")):
                sales_data.append({"ID продажи": row[0], "Название товара": row[1], "Чистая выручка": float(row[2]) if row[2] else 0, "Дата операции": row[3]})
    return sales_data

def parse_services_data(mp_path):
    if not os.path.exists(mp_path): return []
    mp = load_workbook(mp_path, data_only=True)
    sheet = mp.active
    services_data = []
    found_services_header = False
    for row in sheet.iter_rows(values_only=True):
        if not row: continue
        if "Таблица затрат на услуги" in str(row[5]):
            found_services_header = True
            continue
        if found_services_header:
            if row[5] and (str(row[5]).startswith("WB-") or str(row[5]).startswith("SRV") or str(row[5]).startswith("YAN")):
                try:
                    services_data.append({"Услуга": row[6], "Цена": float(str(row[8]).replace(",", ".")) if row[8] else 0.0})
                except (ValueError, IndexError):
                    pass 
    return services_data

def parse_cost_data(mp_path):
    if not os.path.exists(mp_path): return {}
    mp = load_workbook(mp_path, data_only=True)
    sheet = mp.active
    cost_data = {}
    found_cost_header = False
    header_passed = False
    for row in sheet.iter_rows(values_only=True):
        if not row: continue
        if "Таблица себестоимости" in str(row[0]):
            found_cost_header = True
            continue
        if found_cost_header:
            if not header_passed:
                header_passed = True
                continue
            if row[0] and row[1]:
                try:
                    cost_data[row[0]] = float(row[1])
                except (ValueError, TypeError):
                    cost_data[row[0]] = 0
    return cost_data

def build_sales_dict(sales_data):
    sales_dict = {}
    for sale in sales_data:
        product = sale["Название товара"]
        if product not in sales_dict: sales_dict[product] = []
        sales_dict[product].append([sale["Чистая выручка"], sale["Дата операции"]])
    return sales_dict

def build_services_dict(services_data):
    services_dict = {}
    for service in services_data:
        service_name = service["Услуга"]
        if service_name not in services_dict: services_dict[service_name] = []
        services_dict[service_name].append(service["Цена"])
    return services_dict



def calculate_total_revenue(sales_dict):
    total = sum(sale[0] for product, sales_list in sales_dict.items() for sale in sales_list)
    return round(total, 2)

def calculate_total_services_cost(services_dict):
    return round(sum(sum(prices) for prices in services_dict.values()), 2)

def calculate_total_cost_of_goods(sales_dict, cost_dict):
    total = sum(cost_dict.get(product, 0) * len(sales_list) for product, sales_list in sales_dict.items())
    return round(total, 2)

def calculate_avg_cost_per_item(sales_dict, services_dict):
    total_services_cost = sum(sum(prices) for prices in services_dict.values())
    total_items_sold = sum(len(sales) for sales in sales_dict.values())
    return round(total_services_cost / total_items_sold, 2) if total_items_sold > 0 else 0

def get_top_products(sales_dict, top_n=3):
    products_sales = sorted([(product, len(sales)) for product, sales in sales_dict.items()], key=lambda x: x[1], reverse=True)
    return products_sales[:top_n]

def calculate_product_profit(sales_dict, total_services_cost, cost_dict):
    profit_by_product = {}
    total_items_sold = sum(len(sales) for sales in sales_dict.values())
    if total_items_sold == 0: return {}
    for product, sales in sales_dict.items():
        items_count = len(sales)
        revenue = sum(sale[0] for sale in sales)
        cost = cost_dict.get(product, 0) * items_count
        service_cost = (items_count / total_items_sold) * total_services_cost
        profit_by_product[product] = round(revenue - cost - service_cost, 2)
    return profit_by_product

def get_top_profitable_products(profit_dict, top_n=3):
    return sorted(profit_dict.items(), key=lambda x: x[1], reverse=True)[:top_n]

def get_least_sold_products(sales_dict, top_n=3):
    products_sales = sorted([(product, len(sales)) for product, sales in sales_dict.items()], key=lambda x: x[1])
    return products_sales[:top_n]

def get_least_profitable_products(profit_dict, top_n=3):
    return sorted(profit_dict.items(), key=lambda x: x[1])[:top_n]

def calculate_roi(profit, cost_of_goods):
    return round((profit / cost_of_goods) * 100, 2) if cost_of_goods > 0 else 0

def prepare_sales_data(sales_dict):
    if not sales_dict: return pd.DataFrame(columns=['Дата', 'Сумма']).set_index('Дата')
    dates, amounts = [], []
    for product_sales in sales_dict.values():
        for sale in product_sales:
            dates.append(sale[1])
            amounts.append(sale[0])
    df = pd.DataFrame({'Дата': dates, 'Сумма': amounts})
    return df.groupby('Дата')['Сумма'].sum().to_frame()

def plot_sales_dynamics(sales_data, platform_name, graphs_path):
    if sales_data.empty: return
    plt.figure(figsize=(12, 6))
    plt.plot(sales_data.index, sales_data['Сумма'], label='Продажи', marker='o', color='blue')
    mean_value = sales_data['Сумма'].mean()
    plt.axhline(mean_value, color='red', linestyle='--', label=f'Среднее: {mean_value:.2f} руб.')
    plt.title(f'Динамика продаж на {platform_name}')
    plt.xlabel('Дата'); plt.ylabel('Сумма продаж, руб.')
    plt.legend(); plt.grid(True); plt.xticks(rotation=45)
    filename = os.path.join(graphs_path, f'sales_dynamics_{platform_name}.png')
    plt.savefig(filename, bbox_inches='tight', dpi=150)
    plt.close()

def analyze_5day_periods(sales_dict, platform_name, graphs_path):
    if not sales_dict: return pd.DataFrame()
    dates, amounts = [], []
    for product_sales in sales_dict.values():
        for sale in product_sales:
            dates.append(sale[1]); amounts.append(sale[0])
    df = pd.DataFrame({'Дата': dates, 'Сумма': amounts})
    df['ДеньМесяца'] = df['Дата'].dt.day
    df['Период'] = ((df['ДеньМесяца'] - 1) // 5) + 1
    df['Период_Название'] = df['Дата'].dt.strftime('%Y-%m') + '-P' + df['Период'].astype(str)
    period_sales = df.groupby('Период_Название').agg({'Сумма': ['sum', 'mean', 'count'], 'Дата': ['min', 'max']})
    period_sales.columns = ['Сумма продаж', 'Средний чек', 'Количество продаж', 'Начало периода', 'Конец периода']
    period_sales = period_sales.sort_values('Начало периода')
    if period_sales.empty: return pd.DataFrame()
    plt.figure(figsize=(14, 7))
    x_labels = [f"{row['Начало периода'].strftime('%d.%m')}-{row['Конец периода'].strftime('%d.%m')}" for _, row in period_sales.iterrows()]
    plt.step(x_labels, period_sales['Сумма продаж'], where='mid', label='Сумма продаж', linewidth=2.5, color='royalblue')
    plt.scatter(x_labels, period_sales['Сумма продаж'], color='navy', s=50)
    mean_value = period_sales['Сумма продаж'].mean()
    plt.axhline(mean_value, color='red', linestyle='--', label=f'Среднее: {mean_value:,.2f} руб.'.replace(',', ' '))
    plt.title(f'Динамика продаж на {platform_name} (5-дневные периоды)', pad=20)
    plt.xlabel('5-дневные периоды месяца'); plt.ylabel('Сумма продаж, руб.')
    plt.legend(); plt.grid(True, axis='y', linestyle='--', alpha=0.7); plt.xticks(rotation=45, ha='right')
    for i, value in enumerate(period_sales['Сумма продаж']):
        plt.text(i, value + 0.05 * mean_value, f'{value:,.0f}'.replace(',', ' '), ha='center', va='bottom', fontsize=9)
    filename = os.path.join(graphs_path, f'5day_periods_{platform_name}.png')
    plt.tight_layout()
    plt.savefig(filename, bbox_inches='tight', dpi=150)
    plt.close()
    return period_sales




def main(input_path, output_path, relative_json_path):
    try:
      
        cost_dict = parse_cost_data(os.path.join(input_path, "costs.xlsx"))

        platforms_data = {
            "Wildberries": {"sales": [], "services": []},
            "Ozon": {"sales": [], "services": []},
            "Яндекс.Маркет": {"sales": [], "services": []}
        }
        
        platforms_data["Wildberries"]["sales"] = build_sales_dict(parse_sales_data(os.path.join(input_path, "wb_processed_report.xlsx")))
        platforms_data["Wildberries"]["services"] = build_services_dict(parse_services_data(os.path.join(input_path, "wb_processed_report.xlsx")))
        platforms_data["Ozon"]["sales"] = build_sales_dict(parse_sales_data(os.path.join(input_path, "ozon_processed_report.xlsx")))
        platforms_data["Ozon"]["services"] = build_services_dict(parse_services_data(os.path.join(input_path, "ozon_processed_report.xlsx")))
        platforms_data["Яндекс.Маркет"]["sales"] = build_sales_dict(parse_sales_data(os.path.join(input_path, "ym_processed_report.xlsx")))
        platforms_data["Яндекс.Маркет"]["services"] = build_services_dict(parse_services_data(os.path.join(input_path, "ym_processed_report.xlsx")))

        
        graphs_path = os.path.join(output_path, 'graphs')
        os.makedirs(graphs_path, exist_ok=True)

        final_stats = {'platforms': {}, 'general_stats': {}}

        for name, p_data in platforms_data.items():
            sales_dict, services_dict = p_data['sales'], p_data['services']
            
            revenue = calculate_total_revenue(sales_dict)
            services_cost = calculate_total_services_cost(services_dict)
            cost_of_goods = calculate_total_cost_of_goods(sales_dict, cost_dict)
            profit = round(revenue - services_cost - cost_of_goods, 2)
            
            product_profits = calculate_product_profit(sales_dict, services_cost, cost_dict)
            
            
            daily_sales_df = prepare_sales_data(sales_dict)
            plot_sales_dynamics(daily_sales_df, name, graphs_path)
            period_results_df = analyze_5day_periods(sales_dict, name, graphs_path)

            final_stats['platforms'][name] = {
                'revenue': revenue,
                'services_cost': services_cost,
                'profit': profit,
                'avg_cost_per_item': calculate_avg_cost_per_item(sales_dict, services_dict),
                'unique_products': len(sales_dict),
                'top_sales': get_top_products(sales_dict),
                'top_profit': get_top_profitable_products(product_profits),
                'least_sold': get_least_sold_products(sales_dict),
                'worst_profit': get_least_profitable_products(product_profits),
                'roi': calculate_roi(profit, cost_of_goods),
                'graphs': {
                    'sales_dynamics': os.path.join(relative_json_path, 'graphs', f'sales_dynamics_{name}.png').replace('\\', '/'),
                    '5day_periods': os.path.join(relative_json_path, 'graphs', f'5day_periods_{name}.png').replace('\\', '/')
                },
                '5day_periods': period_results_df.to_dict('index') if not period_results_df.empty else {}
            }
        
   
        final_stats['general_stats'] = {
            'total_revenue': round(sum(p['revenue'] for p in final_stats['platforms'].values()), 2),
            'total_profit': round(sum(p['profit'] for p in final_stats['platforms'].values()), 2),
            'total_products': len(set.union(*(set(d['sales'].keys()) for d in platforms_data.values())))
        }
        final_stats['timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        
        def convert(o):
            if isinstance(o, (np.int64, np.int32)): return int(o)
            if isinstance(o, (np.float64, np.float32)): return float(o)
            if isinstance(o, (pd.Timestamp, datetime)): return o.strftime('%Y-%m-%d %H:%M:%S')
            raise TypeError

        json_filename = os.path.join(output_path, 'complete_marketplace_stats.json')
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(final_stats, f, ensure_ascii=False, indent=4, default=convert)

        print(json.dumps({"status": "success", "message": "Marketplace analysis complete.", "json_file": json_filename}))

    except Exception as e:
        print(json.dumps({"status": "error", "message": str(e)}))

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print(json.dumps({"status": "error", "message": "Invalid number of arguments. Expected: input_path, output_path, relative_json_path"}))
    else:
        input_dir = sys.argv[1]
        output_dir = sys.argv[2]
        relative_path = sys.argv[3]
        main(input_dir, output_dir, relative_path)