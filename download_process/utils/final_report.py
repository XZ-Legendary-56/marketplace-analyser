import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook


def save_final_report(sales_df, services_df, ads_df, output_path):
    writer = pd.ExcelWriter(output_path, engine="openpyxl", datetime_format="YYYY-MM-DD")

    # 1. Разбивка по количеству и деление выручки
    if "Кол-во" in sales_df.columns and "Чистая выручка" in sales_df.columns:
        sales_df = sales_df.copy()

        # Приводим "Кол-во" и "Чистая выручка" к числам
        sales_df["Кол-во"] = pd.to_numeric(sales_df["Кол-во"], errors="coerce").fillna(1).astype(int)
        sales_df["Чистая выручка"] = pd.to_numeric(
            sales_df["Чистая выручка"].astype(str).str.replace(",", ".", regex=False), errors="coerce"
        ).fillna(0.0)

        # Делим выручку на количество
        sales_df["Чистая выручка"] = sales_df["Чистая выручка"] / sales_df["Кол-во"]

        # Повторяем строки по количеству
        sales_df = sales_df.loc[sales_df.index.repeat(sales_df["Кол-во"])]
        sales_df = sales_df.reset_index(drop=True)
        sales_df.drop(columns="Кол-во", inplace=True)

        # Присваиваем уникальные ID продажи
        sales_df["ID продажи"] = [f"oz-{i:05}" for i in range(len(sales_df))]

    # 2. Удаляем ненужные столбцы
    for df in [sales_df, services_df, ads_df]:
        for col in ["Схема", "SKU", "Цена продавца", "Розн. цена", "Скидка итого", "Скидка", "Артикул"]:
            if col in df.columns:
                df.drop(columns=col, inplace=True)

    # 3. Приводим даты к формату YYYY-MM-DD
    for df in [sales_df, services_df, ads_df]:
        for date_col in ["Дата операция", "Дата операции"]:
            if date_col in df.columns:
                df[date_col] = pd.to_datetime(df[date_col], errors="coerce").dt.date

    # 4. Экспорт: таблица продаж
    sales_df.to_excel(writer, sheet_name="Итог", startrow=1, startcol=0, index=False)
    ws = writer.book["Итог"]
    ws.cell(row=1, column=1, value="Таблица Продаж")

    # 5. Услуги — справа от продаж
    if not services_df.empty:
        start_col_services = sales_df.shape[1] + 1
        for r_idx, row in enumerate(dataframe_to_rows(services_df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=start_col_services + 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        ws.cell(row=1, column=start_col_services + 1, value="Таблица затрат на услуги")

    # 6. Реклама — ниже
    max_row = max(len(sales_df), len(services_df)) + 5
    if not ads_df.empty:
        ws.cell(row=max_row, column=1, value="Таблица затрат на рекламу")
        for r_idx, row in enumerate(dataframe_to_rows(ads_df, index=False, header=True), start=max_row + 1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # 7. Автоширина
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        col_letter = column_cells[0].column_letter
        ws.column_dimensions[col_letter].width = max_len + 2

    writer.close()